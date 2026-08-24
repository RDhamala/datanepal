"""Nepal National Population and Housing Census 2021 — National Statistics Office.

Source: https://censusresults.nsonepal.gov.np/
Publisher: National Statistics Office (NSO), Government of Nepal
Licence: see catalog/sources/nso-nphc-2021.yml -- NSO states copyright rather
         than an open licence, and the terms recorded there are an explicit
         project decision, not an inference from the site.

Why this source rather than a republisher
----------------------------------------
This is the only route to authoritative local-level statistics for Nepal, and it
reconciles exactly:

    753 local units          28,925,480
     77 institutional rows      239,098
                             -----------
                             29,164,578   = the official 2021 census total

Two alternatives were checked and rejected on evidence. Open Data Nepal
publishes "Total population by sex (country, province, district and local
level)" whose national total is 26,494,504 -- the *2011* census -- under
`year=2016`, with a `parent_code` that is a bare 1..77 sequence rather than a
P-code. Wikidata carries a population statement for only 60 of 753 local units,
mostly 2011.

The INSTITUTIONAL rows
----------------------
Each district lists an INSTITUTIONAL row (sequence code 99) for people not
resident in a household -- barracks, hostels, prisons, hospitals. It is real
population, it is 239,098 people nationally, and it belongs to no local unit.
Dropping it silently would have made every district under-report by the
difference and quietly corrupted every per-capita figure derived from it: the
same shape as the 80Plus bug, and just as invisible to a consistency check.
So it is carried explicitly as its own dimension member and asserted.

Joining to the spine
--------------------
These tables carry no P-codes. Hierarchy is positional -- for Indv01 a sequence
code plus letter case, for Indv17 the column the name sits in. So the join runs
on (district, base name, unit type), where the type comes from the name's own
suffix and is cross-checked against the spine rather than trusted.

That reaches 751 of 753 with zero type disagreements. The remaining two are
handled by a committed crosswalk seed, not by fuzzy matching: NSO writes
"Melanchi" for Melamchi and prefixes "Manang Ngisyang". Guessing at the rest of
Nepali romanisation is exactly what this project refuses to do -- see the note
in transform/seeds/nso_name_fixes.csv.
"""

from __future__ import annotations

import io
import logging
import os
import time
from collections.abc import Iterator
from typing import Any

import dlt
import httpx
import openpyxl

logger = logging.getLogger(__name__)

BASE = "https://censusresults.nsonepal.gov.np/files/province/P{province}/{table}.xlsx"

# nsonepal.gov.np publishes `Crawl-delay: 10`. The census subdomain serves no
# robots.txt at all, but honouring the parent domain's delay is the polite read
# of that, and 21 requests at 10 seconds is a trivial cost for a monthly job.
CRAWL_DELAY_SECONDS = 10

PROVINCES = {
    1: "Koshi",
    2: "Madhesh",
    3: "Bagmati",
    4: "Gandaki",
    5: "Lumbini",
    6: "Karnali",
    7: "Sudur Paschim",
}

# Externally known counts, asserted rather than trusted.
EXPECTED_LOCAL_UNITS = 753
EXPECTED_DISTRICTS = 77
NATIONAL_POPULATION_2021 = 29_164_578

# The suffix names the unit type. Ordered longest-first: "Sub-Metropolitan City"
# ends with "Metropolitan City", and "Upa-Mahanagarpalika" contains
# "Mahanagarpalika" -- the same substring trap the Devanagari suffixes have, and
# the reason CLAUDE.md says to use the P-code type digit rather than a suffix.
# Here there is no type digit, so the suffix is all we have and the resulting
# type is cross-checked against the spine in the transform layer.
#
# "Metropolitian" is not a typo on our side. NSO spells it that way throughout.
# The dimension vocabulary uses 'all' for the total, so NSO's "Total" is mapped
# rather than passed through. A second spelling of the same member would split
# every aggregate silently.
SEX_MEMBERS = {"total": "all", "male": "male", "female": "female"}

TYPE_SUFFIXES: list[tuple[str, str]] = [
    ("Sub-Metropolitian City", "sub_metropolitan"),
    ("Sub-Metropolitan City", "sub_metropolitan"),
    ("Upa-Mahanagarpalika", "sub_metropolitan"),
    ("Upamahanagarpalika", "sub_metropolitan"),
    ("Metropolitian City", "metropolitan"),
    ("Metropolitan City", "metropolitan"),
    ("Mahanagarpalika", "metropolitan"),
    ("Rural Municipality", "rural_municipality"),
    ("Gaunpalika", "rural_municipality"),
    ("Nagarpalika", "municipality"),
    ("Municipality", "municipality"),
]


def _verify() -> str | bool:
    """httpx ignores the CA-bundle environment variables; honour them here."""
    for var in ("REQUESTS_CA_BUNDLE", "SSL_CERT_FILE"):
        path = os.getenv(var)
        if path and os.path.exists(path):
            return path
    return True


def _split_type(name: str) -> tuple[str, str | None]:
    """Split a local-unit name into its base name and unit type."""
    cleaned = " ".join(str(name).split())
    for suffix, kind in TYPE_SUFFIXES:
        if cleaned.lower().endswith(suffix.lower()):
            return cleaned[: -len(suffix)].strip(), kind
    return cleaned, None


def _fetch(province: int, table: str, session: httpx.Client) -> openpyxl.Workbook:
    url = BASE.format(province=province, table=table)
    response = session.get(url)
    response.raise_for_status()
    # read_only would be faster but reports merged cells inconsistently, and the
    # area hierarchy in these files is expressed through merged cells.
    return openpyxl.load_workbook(io.BytesIO(response.content), data_only=True)


def _cell_text(value: Any) -> str:
    return "" if value is None else " ".join(str(value).split())


def _as_int(value: Any) -> int | None:
    if value is None or value == "":
        return None
    try:
        return int(float(value))
    except (TypeError, ValueError):
        return None


# --------------------------------------------------------------- population

@dlt.resource(name="census_population", write_disposition="replace", primary_key="row_id")
def census_population() -> Iterator[dict[str, Any]]:
    """Table Indv01: households and population by sex, every admin level.

    Layout is a sequence code in column A and a single name column in B:

        code 0  + UPPER CASE name  -> province heading
        code 0  + Title Case name  -> district heading
        code 1..N                  -> local unit within the current district
        code 99                    -> INSTITUTIONAL, not a place
    """
    counts = {"province": 0, "district": 0, "local": 0, "institutional": 0}
    national: int | None = None
    local_total = 0
    institutional_total = 0

    with httpx.Client(
        timeout=120,
        follow_redirects=True,
        verify=_verify(),
        headers={"User-Agent": "DataNepalBot/1.0 (+https://datanepal.org)"},
    ) as session:
        for province_no, province_name in PROVINCES.items():
            workbook = _fetch(province_no, "Indv01-PopulationBySex", session)
            sheet = workbook[workbook.sheetnames[0]]
            district: str | None = None

            for row in sheet.iter_rows(min_row=6, values_only=True):
                name = _cell_text(row[1])
                if not name:
                    continue
                code = _as_int(row[0])
                households, total, male, female = (
                    _as_int(row[2]),
                    _as_int(row[3]),
                    _as_int(row[4]),
                    _as_int(row[5]),
                )

                if name == "Nepal":
                    # Present in every workbook; emit once. Publishing it is not
                    # decoration -- assert_population_hierarchy_sums compares
                    # each level against the national figure, and without a
                    # country row that test passes vacuously.
                    if national is None:
                        national = total
                        yield {
                            "row_id": "country|Nepal",
                            "level": "country",
                            "province_name": None,
                            "district_name": None,
                            "raw_name": "Nepal",
                            "base_name": "Nepal",
                            "unit_type": "country",
                            "households": households,
                            "population_total": total,
                            "population_male": male,
                            "population_female": female,
                        }
                    continue
                if code is None:
                    continue

                if code == 0:
                    if name.isupper():
                        counts["province"] += 1
                        district = None
                        yield {
                            "row_id": f"prov|{province_name}",
                            "level": "province",
                            "province_name": province_name,
                            "district_name": None,
                            "raw_name": province_name,
                            "base_name": province_name,
                            "unit_type": "province",
                            "households": households,
                            "population_total": total,
                            "population_male": male,
                            "population_female": female,
                        }
                    else:
                        district = name
                        counts["district"] += 1
                        yield {
                            "row_id": f"dist|{province_name}|{name}",
                            "level": "district",
                            "province_name": province_name,
                            "district_name": name,
                            "raw_name": name,
                            "base_name": name,
                            "unit_type": "district",
                            "households": households,
                            "population_total": total,
                            "population_male": male,
                            "population_female": female,
                        }
                    continue

                if name.upper() == "INSTITUTIONAL" or code == 99:
                    counts["institutional"] += 1
                    institutional_total += total or 0
                    # Attached to the district, because that is the only level
                    # NSO reports it at. Carried, never silently dropped.
                    yield {
                        "row_id": f"inst|{province_name}|{district}",
                        "level": "institutional",
                        "province_name": province_name,
                        "district_name": district,
                        "raw_name": "INSTITUTIONAL",
                        "base_name": None,
                        "unit_type": None,
                        "households": households,
                        "population_total": total,
                        "population_male": male,
                        "population_female": female,
                    }
                    continue

                base, unit_type = _split_type(name)
                counts["local"] += 1
                local_total += total or 0
                yield {
                    "row_id": f"local|{province_name}|{district}|{base}",
                    "level": "local",
                    "province_name": province_name,
                    "district_name": district,
                    "raw_name": name,
                    "base_name": base,
                    "unit_type": unit_type,
                    "households": households,
                    "population_total": total,
                    "population_male": male,
                    "population_female": female,
                }

            time.sleep(CRAWL_DELAY_SECONDS)

    logger.info(
        "census population: %d provinces, %d districts, %d local units, %d institutional rows",
        counts["province"], counts["district"], counts["local"], counts["institutional"],
    )

    problems: list[str] = []
    if counts["local"] != EXPECTED_LOCAL_UNITS:
        problems.append(f"{counts['local']} local units (expected {EXPECTED_LOCAL_UNITS})")
    if counts["district"] != EXPECTED_DISTRICTS:
        problems.append(f"{counts['district']} districts (expected {EXPECTED_DISTRICTS})")
    if counts["province"] != len(PROVINCES):
        problems.append(f"{counts['province']} provinces (expected {len(PROVINCES)})")
    if national != NATIONAL_POPULATION_2021:
        problems.append(f"national row {national} (expected {NATIONAL_POPULATION_2021})")
    # The reconciliation that catches a partial load, which no row count would.
    if local_total + institutional_total != NATIONAL_POPULATION_2021:
        problems.append(
            f"local {local_total:,} + institutional {institutional_total:,} "
            f"= {local_total + institutional_total:,}, expected {NATIONAL_POPULATION_2021:,}"
        )
    if problems:
        raise ValueError("NSO census population load is wrong: " + "; ".join(problems))


# ----------------------------------------------------------------- literacy

@dlt.resource(name="census_literacy", write_disposition="replace", primary_key="row_id")
def census_literacy() -> Iterator[dict[str, Any]]:
    """Table Indv17: literacy of the population aged 5 and over.

    Layout is positional rather than coded -- the column a name sits in gives its
    level, and the rows beneath it are sex then age:

        column A -> province      column D -> sex (Total/Male/Female)
        column B -> district      column E -> age ("Total", "05 Year", ...)
        column C -> local unit

    Only the age "Total" row is emitted. The single-year detail is 13,000 rows
    per province and answers nothing a place profile asks.

    The denominator is population aged 5 and over, not total population. Dividing
    literate people by total population would understate literacy everywhere by
    the size of the under-five cohort, so the base is carried as its own column
    and the rate is computed downstream from it.
    """
    seen_local = 0
    seen_district = 0

    with httpx.Client(
        timeout=180,
        follow_redirects=True,
        verify=_verify(),
        headers={"User-Agent": "DataNepalBot/1.0 (+https://datanepal.org)"},
    ) as session:
        for province_no, province_name in PROVINCES.items():
            workbook = _fetch(province_no, "Indv17-PopulationByLiteracyStatus", session)
            sheet = workbook[workbook.sheetnames[0]]

            district: str | None = None
            local: str | None = None
            level = "province"
            sex: str | None = None

            for row in sheet.iter_rows(min_row=6, values_only=True):
                province_cell = _cell_text(row[0])
                district_cell = _cell_text(row[1])
                local_cell = _cell_text(row[2])
                sex_cell = _cell_text(row[3])
                age_cell = _cell_text(row[4])

                # A name in a column resets everything below it.
                if province_cell:
                    level, district, local = "province", None, None
                    continue
                if district_cell:
                    level, district, local = "district", district_cell, None
                    seen_district += 1
                    continue
                if local_cell:
                    # INSTITUTIONAL sits in the local-unit column here, unlike
                    # Indv01 where it has its own sequence code. Unhandled, it
                    # inflated Koshi from 137 local units to 151 and would have
                    # been counted as a place.
                    if local_cell.upper() == "INSTITUTIONAL":
                        level, local = "institutional", None
                    else:
                        level, local = "local", local_cell
                        seen_local += 1
                    continue
                if sex_cell:
                    sex = SEX_MEMBERS.get(sex_cell.lower())
                    continue

                if age_cell != "Total" or sex is None:
                    continue

                base, unit_type = (
                    _split_type(local) if level == "local" and local else (None, None)
                )
                name = {
                    "province": province_name,
                    "district": district,
                    "local": local,
                    "institutional": "INSTITUTIONAL",
                }[level]
                if name is None:
                    continue

                yield {
                    "row_id": f"{level}|{province_name}|{district}|{local}|{sex}",
                    "level": level,
                    "province_name": province_name,
                    "district_name": district,
                    "raw_name": name,
                    # Institutional rows carry no place name of their own, so
                    # base_name stays null in both tables and the crosswalk
                    # holds one row per district rather than two.
                    "base_name": (
                        base if level == "local"
                        else None if level == "institutional"
                        else name
                    ),
                    # Null for institutional in both tables, matching base_name.
                    # When these two disagreed, the crosswalk held 154 rows
                    # instead of 77 and every institutional observation was
                    # emitted twice -- 308 duplicate rows, caught by the
                    # uniqueness test rather than by anything visible.
                    "unit_type": (
                        unit_type if level == "local"
                        else None if level == "institutional"
                        else level
                    ),
                    "sex": sex,
                    # Population aged 5+, the denominator for every rate here.
                    "population_5plus": _as_int(row[5]),
                    "can_read_and_write": _as_int(row[6]),
                    "can_read_only": _as_int(row[7]),
                    "cannot_read_or_write": _as_int(row[8]),
                    "literacy_not_stated": _as_int(row[9]),
                }

            time.sleep(CRAWL_DELAY_SECONDS)

    # Three sex rows per area, so the local count is three times the units.
    logger.info(
        "census literacy: %d local rows, %d district rows", seen_local, seen_district
    )
    if seen_local != EXPECTED_LOCAL_UNITS:
        raise ValueError(
            f"NSO literacy load saw {seen_local} local units, expected {EXPECTED_LOCAL_UNITS}"
        )
    if seen_district != EXPECTED_DISTRICTS:
        raise ValueError(
            f"NSO literacy load saw {seen_district} districts, expected {EXPECTED_DISTRICTS}"
        )


@dlt.source(name="nso_census")
def nso_census_source():
    return [census_population(), census_literacy()]
